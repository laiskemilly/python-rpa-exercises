from openpyxl import load_workbook
import os
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

caminho_arquivo =  'InserirDados_PintarCelulas.xlsx'
planilha_aberta = load_workbook(filename=caminho_arquivo)

sheet_selecionada = planilha_aberta['Aluno']

dados_tabela = [
    ['Nome', 'Idade'],
    ['Ana', 18],
    ['João', 16],
    ['Gabriel', 26],
    ['Julia', 25],
    ['Maria', 29]
]


for linha_planilha in dados_tabela:
    sheet_selecionada.append(linha_planilha)

cor_titulo = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
cor_celula = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type='solid')

sheet_selecionada["A1"].fill = cor_titulo 
sheet_selecionada["B1"].fill = cor_titulo 

for linha in range (2, len(sheet_selecionada['A'])+ 1):
    celula_colunaA= "A" + str(linha)
    celula_colunaB= "B" + str(linha)

    sheet_selecionada[celula_colunaA].fill = cor_celula
    sheet_selecionada[celula_colunaB].fill = cor_celula

planilha_aberta.save(filename=caminho_arquivo)
os.startfile(caminho_arquivo)